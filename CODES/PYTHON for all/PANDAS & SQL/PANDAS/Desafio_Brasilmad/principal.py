import pandas as pd
import os
import openpyxl as op

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# os.path.abspath(__file__)

wb=op.load_workbook('TesteBrasilmadVBA.xlsm')
db=pd.read_excel('TesteBrasilmadVBA.xlsm')
a=wb['BD COMPRAS']
sheet0=wb[wb.sheetnames[0]]
sheet1=wb[wb.sheetnames[1]]
sheet2=wb[wb.sheetnames[2]]

# sheet.cell(row=2, column=3).value = 95
DataNF1 = sheet0['C18'].value
DataNF1 = sheet0['E18'].value

NumNF = sheet0['C19'].value
SITs = sheet0['C20'].value
fornecedor=sheet0['C21'].value

max_row= sheet1.max_row
max_column= sheet1.max_column

def busca():
    for row in range(7,max_row):
        sheet1.cell(row=row, column=5).value
        if sheet1.cell(row=row, column=5).value == sheet0['C20'].value:
            sheet0.cell(row=row+11, column=12).value = sheet1.cell(row=row, column=5).value


busca()
# wb.save('Teste_B.xlsm')


