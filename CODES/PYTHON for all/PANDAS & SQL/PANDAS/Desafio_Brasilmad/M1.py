import pandas as pd
import os
import openpyxl as op

os.chdir('C:\\CODES\Pessoal Project\\Macro Excel com Python')
wb=op.load_workbook('TesteBrasilmadVBA.xlsx')
db=pd.read_excel('TesteBrasilmadVBA.xlsx')
a=wb['BD COMPRAS']
sheet0=wb[wb.sheetnames[0]]
sheet1=wb[wb.sheetnames[1]]
sheet2=wb[wb.sheetnames[2]]


print(sheet0)

from numpy.random import randn
df = pd.DataFrame (randn(5,4)*10, index=["A", "B", "C", "D", "E"], columns="W X Y Z".split())
print(df)

print(df[['X']])   # printa no formato data frame ao invès de string

df. drop("y",axis=1, inplace=True)  # axis=1 = COLUNAS  , axis=0  = LINHAS    inplace=True faz alterações no df original ao invéz da cópia

b =df.iloc[0,2]

a = df.loc['A',"B"]
print(a)
print(b)