import pandas as pd
import os
import openpyxl as op

def busca_pandas():
    # Lê a planilha em um DataFrame
    df = pd.read_excel('TesteBrasilmadVBA.xlsx', sheet_name='BD COMPRAS', skiprows=6)
    
    # Filtra as linhas onde a coluna 5 (E) corresponde ao valor desejado
    valor_filtro = sheet0['C20'].value
    linhas_filtradas = df[df.iloc[:, 4] == valor_filtro]
    
    # Define a linha de início para escrita dos resultados
    linha_inicial = 30  # Ajuste este número para a linha onde você quer começar a escrita
    
    # Escreve as linhas filtradas na planilha de destino
    for i, (idx, row) in enumerate(linhas_filtradas.iterrows()):
        linha_atual = linha_inicial + i
        for col_idx, valor in enumerate(row.values):
            # Verifica se a célula é mesclada antes de tentar escrever
            cell = sheet0.cell(row=linha_atual, column=col_idx+1)
            if not isinstance(cell, op.cell.MergedCell):
                cell.value = valor

    return linhas_filtradas

# Código principal
os.chdir(r'C:\\CODES\\PYTHON for all\\PANDAS')  # Usando r-string para evitar escape sequences

wb = op.load_workbook('TesteBrasilmadVBA.xlsx')
sheet0 = wb[wb.sheetnames[0]]
sheet1 = wb[wb.sheetnames[1]]
sheet2 = wb[wb.sheetnames[2]]

# Suas variáveis existentes
DataNF1 = sheet0['C18'].value
NumNF = sheet0['C19'].value
SITs = sheet0['C20'].value
fornecedor = sheet0['C21'].value

# Chama a função de busca
resultados = busca_pandas()

wb.save('TesteBrasilmadVBA.xlsx')