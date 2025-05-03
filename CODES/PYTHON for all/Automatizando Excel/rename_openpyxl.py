import os
from openpyxl import load_workbook
os.chdir('C:/PYTHON for All/Automatizando Excel/')

wb= load_workbook('C:/PYTHON for All/Automatizando Excel/exemplo.xlsx')


sheetnames = wb.sheetnames
# wb.get_sheet_by_name   é agora wb.worksheets
local_planilia=wb.worksheets[0]['A2'].value   #valor escolhido por célula


sheet=wb['Sheet1']['B3'].value
sheet=wb['Sheet1']
sheet.cell(row=2, column=2).value

sheet.max_column
sheet.max_row

# pega as colunas
for i in range(0,sheet.max_row):
    print(sheet.cell(row=i+1,column=2).value)
# pega as linhas
for i in range(0,sheet.max_column):
    print(sheet.cell(row=3,column=i+1).value)

sheet.cell(row=2, column=3).value = 95
wb.save('exemplo.xlsx')

# AGRUPAMENTO
def AGRUPAMENTO_FUNÇÃO_DESNECESSÁRIA(): 
    sheet.merge_cells("A1:D1")
    sheet.unmerge_cells("A1:D1")

sheet.insert_rows(4)
sheet.insert_cols(8)
sheet.delete_cols(2,5) #DELETA DA B:F

# ADIÇÃO DE IMAGENS
from openpyxl.drawing.image  import Image
img=('C:/PYTHON for All/Automatizando Excel/Automatizando Excel/catlogo.png')

# desatualizado, confira em openpyxl.readthedocs.io/en/stable/
sheet.add_image(img,"A8")



# print(sheetnames)
print(sheet)