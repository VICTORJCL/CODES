from openpyxl import Workbook, load_workbook
import os
os.chdir("C:\\CODES\\PYTHON for all\\BASIC TESTS\\excel")
wb=load_workbook("C:\\CODES\\PYTHON for all\\BASIC TESTS\\excel\\dados2 power bi.xlsx")

# verificando abas e pegando abas  'sheets'
print(wb.sheetnames)
# sheet_base=wb['Base']
# sheet_base=wb.sheetnames[1]
sheet_base=wb[wb.sheetnames[1]]
print(type(sheet_base))     #<class 'openpyxl.worksheet.worksheet.Worksheet'>
print(sheet_base)

# verificar o máximo de linhas  e colunas
print(sheet_base.max_row)
print(sheet_base.max_column)
# alterar valores
sheet_base.cell(row=2, column=1).value = 'Lira'

# sheet_base['A4']  célula
sheet_base['A4'].value ='Andrei'

# rowmax
# print(type(content))
sheet_base['I2'].value = 35
wb.save('dados2 power bi.xlsx')
sheet_base




# rag = [sheet_base.cell(row=i, column=8).value for i in range(2, sheet_base.max_row+1)]
# print(type(rag[0]))
# print(rag)

for i in range(2,sheet_base.max_row+1):
    sheet_base.cell(row=i, column=10).value =  f"=(H{i}*0.6)"



for i in range(2,sheet_base.max_row+1):
    sheet_base.cell(row=i, column=9).value = sheet_base.cell(row=i, column=8).value * 0.6
wb.save('dados2 power bi.xlsx')
print(sheet_base['J2'].value)
print(sheet_base['J2'])